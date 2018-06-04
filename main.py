import params
import importlib
import numpy as np
from tensorflow.python.platform import gfile
from funcData import *
from evaluate import evaluate
from tqdm import tqdm
import funcCustom
import funcQuantize
import funcSummary


FLAGS = tf.app.flags.FLAGS
# tf.logging.set_verbosity(FLAGS.log)
if FLAGS.dataset=='cifar10':
    steps=round(50000/FLAGS.batch_size)*20
else:
    steps=round(55000/FLAGS.batch_size)*20
##LR을 decay시켜주는 함수
def learning_rate_decay_fn(learning_rate, global_step,decay_steps=steps):
    print("learning_rate_decay_fn is executed!")
    return tf.train.exponential_decay(
      learning_rate,
      global_step,
      decay_steps=decay_steps,
      decay_rate=0.8,
      staircase=True)
def quantizeGrads(Grad_and_vars,target_level=FLAGS.W_target_level):
    if target_level <= 256*256:
        grads = []
        for grad_and_vars in Grad_and_vars:
            grads.append([funcQuantize.quantize_G(grad_and_vars[0], target_level), grad_and_vars[1]])
        return grads
    return Grad_and_vars

## model을 data로 training 시켜주는 함수

def train(model, data,
          batch_size=128,
          learning_rate=FLAGS.learning_rate,
          Weight_decay=FLAGS.Weight_decay,
          log_dir='./log',
          checkpoint_dir='./checkpoint',
          num_epochs=-1):
    with tf.name_scope('data'):
        x, yt = data.next_batch(batch_size)
    global_step =  tf.get_variable('global_step', shape=[],dtype=tf.int64,initializer=tf.constant_initializer(0),
                         trainable=False)
    tf.add_to_collection("Step",global_step)  #Evaluate에서 Drift효과 끄기 위해 구분점역할을 한다.
    y = model(x, is_training=True)
    # Define loss and optimizer
    with tf.name_scope('objective'):
        yt_one=tf.one_hot(yt,10)
        loss = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits_v2(labels=yt_one, logits=y), name="loss")
        if Weight_decay:
            loss=loss+tf.reduce_sum(params.WEIGHT_DECAY_FACTOR * tf.stack([tf.nn.l2_loss(i) for i in tf.get_collection('Original_Weight', scope='L')]))
        accuracy=tf.reduce_mean(tf.cast(tf.equal(tf.argmax(yt_one,1), tf.argmax(y, axis=1)),dtype=tf.float32),name="accuracy")
    vars_train=tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L22')\
               +tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L23')\
               +tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L24')\
               +tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES,scope='L25') if FLAGS.Fine_tuning else None
    # # * 이런 식으로 gradient  뽑아서  수정가능
    optimizer=tf.train.GradientDescentOptimizer(1)
    grads = optimizer.compute_gradients(loss)
    gradTrainBatch_quantize = quantizeGrads(grads,FLAGS.W_target_level)
    opt = optimizer.apply_gradients(gradTrainBatch_quantize, global_step=global_step)
    if FLAGS.W_target_level>2**10:
        opt = tf.contrib.layers.optimize_loss(loss, global_step, learning_rate, optimizer='Adam',
                                              gradient_noise_scale=None, gradient_multipliers=None,
                                              clip_gradients=None, # moving_average_decay=0.9,
                                               update_ops=None, variables=vars_train, name=None)

    print("Definite Moving Average...")
    ema = tf.train.ExponentialMovingAverage(params.MOVING_AVERAGE_DECAY, global_step, name='average')
    ema_op = ema.apply([loss, accuracy]+tf.trainable_variables())
    tf.add_to_collection(tf.GraphKeys.UPDATE_OPS, ema_op)
    loss_avg = ema.average(loss)
    tf.summary.scalar('loss/training', loss_avg)
    accuracy_avg = ema.average(accuracy)
    tf.summary.scalar('accuracy/training', accuracy_avg)
    check_loss = tf.check_numerics(loss, 'model diverged: loss->nan')
    tf.add_to_collection(tf.GraphKeys.UPDATE_OPS, check_loss)
    list_W = tf.get_collection('Original_Weight', scope='L')
    list_Wbin = tf.get_collection('Binarized_Weight', scope='L')
    list_Wfluc = tf.get_collection('Fluctuated_Weight', scope='L')
    list_Wread = tf.get_collection('Read_Weight',scope='L')
    list_Wprop = tf.get_collection('Propagated_Weight',scope='L')
    list_Drift_step = tf.get_collection('Drift_step', scope='L')
    list_Drift_value = tf.get_collection('Drift_value', scope='L')
    list_pre_Wbin = tf.get_collection('pre_Wbin', scope='L')
    list_pre_Wfluc = tf.get_collection('pre_Wfluc', scope='L')
    list_pre_Wbin_op = tf.get_collection('pre_Wbin_update_op', scope='L')
    list_pre_Wfluc_op = tf.get_collection('pre_Wfluc_update_op', scope='L')
    clip_op_list=[]
    with tf.control_dependencies([opt]):
        for ww in list_W:
            clip_op=tf.assign(ww, funcQuantize.clip(ww, FLAGS.W_target_level))
            clip_op_list+=[clip_op]
    updates_collection = tf.get_collection(tf.GraphKeys.UPDATE_OPS)
    # updates_collection은 여러 operation으로 이루어진 리스트고, train_op는 그 operation들을 실행하는 operation이다.
    with tf.control_dependencies(clip_op_list):
        train_op = tf.group(*updates_collection)
    print("Make summary for writer...")
    if FLAGS.summary:
        funcSummary.add_summaries_scalar([accuracy,loss])
        funcSummary.add_summaries_weight(list_W)
        # funcSummary.add_summaries_weight(list_Wbin)
        # funcSummary.add_summaries_weight(list_Wfluc)
        funcSummary.add_summaries_weight(list_Wread)
        funcSummary.add_summaries_weight(list_Wprop)
    summary_op = tf.summary.merge_all()
    print("Open Session...")
    # Configure options for session
    gpu_options = tf.GPUOptions(allow_growth=True,per_process_gpu_memory_fraction=0.9)
    sess = tf.InteractiveSession(
        config=tf.ConfigProto(
            log_device_placement=False,
            allow_soft_placement=True, gpu_options=gpu_options))
    sess.run(tf.global_variables_initializer())
    for W_ori in list_W:
        W_ori_new=sess.run(tf.assign(W_ori, funcQuantize.quantize_W(W_ori, FLAGS.W_target_level)))
    if FLAGS.Load_checkpoint:
        call_list=tf.get_collection(tf.GraphKeys.VARIABLES)
        print("******We will restore:",call_list)
        saver = tf.train.Saver(max_to_keep=1,var_list=call_list)
        ckpt = tf.train.get_checkpoint_state(params.load_checkpoint_path)
        if ckpt and ckpt.model_checkpoint_path:
            # Restores from checkpoint
            saver.restore(sess, ckpt.model_checkpoint_path)
        else:
            print('No checkpoint file found')
    saver = tf.train.Saver(max_to_keep=1)
    # saver_best= tf.train.Saver(max_to_keep=1)
    summary_writer = tf.summary.FileWriter(log_dir, graph=sess.graph)
    coord = tf.train.Coordinator()
    threads = tf.train.start_queue_runners(sess=sess, coord=coord)
    best_acc = 0
    best_loss = 0
    patience = 0
    num_batches = int(data.size[0] / batch_size)
    print("Check the collections...")
    print("list_W:\n",list_W,'\nNum:',len(list_W))
    print("Activations:\n", tf.get_collection(tf.GraphKeys.ACTIVATIONS), '\nNum:', len(tf.get_collection(tf.GraphKeys.ACTIVATIONS)))

    file = open(FLAGS.checkpoint_dir + "/save_model.py", "a")
    funcCustom.magic_print(
        'We start training..num of trainable param: %d' % funcCustom.count_params(tf.trainable_variables()),file=file)

    for i in range(num_epochs):
        if len(params.LR_schedule)/2:
            if i == params.LR_schedule[0]:
                params.LR_schedule.pop(0)
                LR_new = params.LR_schedule.pop(0)
                if LR_new == 0:
                    print('Optimization Ended!')
                    exit(0)
                LR_old = sess.run(params.LR)
                sess.run(params.LR.assign(LR_new))
                print('lr: %f -> %f' % (LR_old, LR_new))
        funcCustom.magic_print('Started epoch %d' % (i + 1), file=file)
        count_num=np.array([0,0,0,0,0,0,0,0,0,0])
        for j in tqdm(range(num_batches)):
            pre_iter_info=sess.run(list_Wbin+list_Wfluc)
            list_run = sess.run([train_op, loss]+[y,yt])
            unique_elements,elements_counts=np.unique(list_run[-1],return_counts=True)
            num_set=dict(zip(unique_elements,elements_counts))
            #ii라는 숫자가 dictionary에 들어있다면 카운트에 더해준다.
            for ii in range(10):
                if num_set.__contains__(ii):
                    count_num[ii]=count_num[ii]+num_set[ii]
            if FLAGS.Inter_variation_options:
                for index, value in enumerate(pre_iter_info[0:len(list_Wbin)]):
                    sess.run(list_pre_Wbin_op[index],{list_pre_Wbin[index]:value})
                for index, value in enumerate(pre_iter_info[len(list_Wbin):len(list_Wbin + list_Wfluc)]):
                    sess.run(list_pre_Wfluc_op[index],{list_pre_Wfluc[index]:value})
            if j%100==0:
                summary_writer.add_summary(sess.run(summary_op), global_step=sess.run(global_step))
        step, acc_value, loss_value, summary = sess.run([global_step, accuracy_avg, loss_avg, summary_op])
        # step, acc_value, loss_value = sess.run([global_step, accuracy_avg, loss_avg])
        funcCustom.magic_print(
            ["%d : " % i + str(count_num[i]) for i in range(10)], " Totral num: ", count_num.sum(),file=file)
        funcCustom.magic_print('Training - Accuracy: %.3f' % acc_value, '  Loss:%.3f' % loss_value, file=file)
        saver.save(sess, save_path=checkpoint_dir + '/model.ckpt', global_step=global_step)
        test_acc, test_loss = evaluate(model, FLAGS.dataset,checkpoint_dir=checkpoint_dir)# log_dir=log_dir)
        funcCustom.magic_print('Test     - Accuracy: %.3f' % test_acc, '  Loss:%.3f' % test_loss, file=file)
        if best_acc<test_acc:
            best_acc=test_acc
            best_loss=test_loss
            patience=0
            # saver_best.save(sess, save_path=checkpoint_dir + '/best_model.ckpt', global_step=global_step,latest_filename="best_checkpoint")
        elif best_acc >= test_acc:
            patience += 1
            if patience > 20:
                funcCustom.magic_print("Stop this training at epoch" + str(i + 1) + ", because accuracy may be saturated", file=file)
                from openpyxl import Workbook
                from openpyxl import load_workbook
                file_name = 'Accuracy_log'+FLAGS.daystr
                if gfile.Exists(file_name+'.xlsx'):
                    wb=load_workbook(filename=file_name+'.xlsx')
                    ws=wb[file_name]
                else:
                    wb=Workbook()
                    ws=wb.active
                    ws.title=file_name
                row=13 if FLAGS.dataset=='cifar10' else 2  #cifar10:2-10   MNIST:11-19
                row=row+13*(FLAGS.row_choice+FLAGS.choice)
                ws.cell(column=FLAGS.col_choice + 1, row=row , value=FLAGS.W_target_level)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 1, value=FLAGS.Wq_target_level)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 2, value=best_acc)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 3, value=best_loss)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 4, value=test_acc)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 5, value=test_loss)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 6, value=acc_value)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 7, value=loss_value)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 8, value=i)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 9, value=FLAGS.Inter_variation_options)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 10, value=FLAGS.Intra_variation_options)
                ws.cell(column=FLAGS.col_choice + 1, row=row + 11, value="Set"+str(FLAGS.choice+1)+"_"+FLAGS.model)
                wb.save(filename=file_name+'.xlsx')
                break
        funcCustom.magic_print('Best     - Accuracy: %.3f(patience=%d)' % (best_acc, patience), file=file)
        summary_out = tf.Summary()
        summary_out.ParseFromString(summary)
        summary_out.value.add(tag='accuracy/test', simple_value=test_acc)
        summary_out.value.add(tag='loss/test', simple_value=test_loss)
        summary_writer.add_summary(summary_out, step)
        summary_writer.flush()


    # When done, ask the threads to stop.
    file.close()
    coord.request_stop()
    coord.join(threads)
    coord.clear_stop()
    summary_writer.close()
"""
설명2:
1)what is the argv
Argv in Python
The list of command line arguments passed to a Python script. argv[0] is the script name (it is operating system 
dependent whether this is a full pathname or not). If the command was executed using the -c command line option 
to the interpreter, argv[0] is set to the string '-c'.

지금은 FLAGS가 글로벌하게 선언이 되어있어서 argv가 전달된다는게 큰의미는 없다.
(전달안되도 어차피 글로벌이라 그냥 사용가능.tf.app.run()은 그냥 one line fast argument parser로 생각하면 될 듯 하다.)
"""

def main(argv=None):  # pylint: disable=unused-argument
    """
    설명3:
    1) gfile은 약간 폴더,파일쪽 다루는 패키지인가보다, 만약 checkpoint_dir에서 지정 된 폴더가 없으면 그걸 만들어서
    2) os.path.join은 단순히 경로 만들어주는 함수다, 저절로 / 를 추가해주는게 편리한 점
    3) assert FLAGS.model로 명시한 모델이 있으면 통과 없으면 뒤에 있는 오류메시지 발생
    4) 해당 파이썬 파일을(인풋1) 인풋2로 복사한다.
    5) 해당 모델을 import한다, importlib.import_module()함수는 코드 과정 중에 패키지를 import 할 때 쓰는 듯 하다.
    6) data = get_data_provider(FLAGS.dataset, training=True) 에서 training에 따라 trainset 혹은 testset을 불러온다.
    7) 위에서 정의한 train함수를 실행 - train을 진행하기 전에 data.py파일을 살펴보자
    """
    if not gfile.Exists(FLAGS.checkpoint_dir):
        # gfile.DeleteRecursively(FLAGS.checkpoint_dir)
        gfile.MakeDirs(FLAGS.checkpoint_dir)
        model_file =( './models/'+FLAGS.model+'.py')
        print(model_file)
        assert gfile.Exists(model_file), 'no model file named: ' + model_file
        gfile.Copy(model_file, FLAGS.checkpoint_dir + '/save_model.py')
        gfile.Copy('./main.py',FLAGS.checkpoint_dir+'/save_main.py')
        gfile.Copy('./evaluate.py',FLAGS.checkpoint_dir+'/save_evaluate.py')
        gfile.Copy('./params.py', FLAGS.checkpoint_dir + '/save_params.py')
        gfile.Copy('./funcData.py', FLAGS.checkpoint_dir + '/save_funcData.py')
        gfile.Copy('./funcCustom.py', FLAGS.checkpoint_dir + '/save_funcCustom.py')
        gfile.Copy('./funcLayer.py', FLAGS.checkpoint_dir + '/save_funcLayer.py')
        gfile.Copy('./funcQuantize.py', FLAGS.checkpoint_dir + '/save_funcQuantize.py')
        gfile.Copy('./funcSummary.py', FLAGS.checkpoint_dir + '/save_funcSummary.py')
    m = importlib.import_module('models.' + FLAGS.model)
    data = get_data_provider(FLAGS.dataset, training=True)

    train(m.model, data,
          batch_size=FLAGS.batch_size,
          checkpoint_dir=FLAGS.checkpoint_dir,
          log_dir=FLAGS.log_dir,
          num_epochs=FLAGS.num_epochs)

"""
설명1:여기서부터 설명한다,일단 이 코드를 실행하면 위에서부터 라인바이라인 실행이 된다.
다만 함수는 라인바이라인 실행 되는 것이 아니라 함수가 있다는 것을 알려줄 수 있는 '선언'만 하고 넘어가게 된다.
그러면 마지막에 실행되는 것이 아래의 조건문이다.
모든 파일은 실행 될 때 __name__이 부여되는 듯 하다. __main__이라는건 이 파일이! 모듈 같이 간접적으로 실행되었다는게 아니라
직접적으로 실행되었다는 뜻이다, 즉 아래의 코드는 이 파일이 직접 실행 된건지, 아니면 다른 코드에 종속적으로 실행된건지를 판단한다.
-여기서 종속적으로 실행되면 __name__=? 인지는 아직 모른다.
이걸 판단해서 직접적으로 실행된거면 tf.app.run()을 실행한다, 아니면 아무것도 실행안된다, 즉 아무것도 안된다.
이는 이 파일이 부정할 수 없는 메인파일이라는 것을 말해준다.

여기서 처음에 들었던 의문-tf.app.run()은 argv를 받아서 main()이라는 이름의 함수(default)에 전달해서 실행시킨다고 하는데
왜 바로 아래 판단문에 그 함수의 내용을 쓰지 않는걸까? 굳이 한번의 매개를 거치는 이유가 무엇인가?
->답변: 우리는 FLAGS를 이용해서 커맨드창에서 파라미터설정을 입력 받는다, 그리고 그걸 wraping해서 함수에 쓰는거다.
tf.app.run()은 그걸 wraping해주는 좋은 함수이고, 그걸 main()함수에 전달하도록 되어있다. 즉 tf.app.run()함수 자체가 main()함수를
필요로 하는 것이다.
#Runs the program with an optional 'main' function and 'argv' list.
->그렇다면 질문은 tf.app.run()가 왜 이렇게 쓰이냐!로 되는데 그건 나중에 생각해보자.
추측: 만약 이 함수를 안쓰면 우리는 string인 cmd 명령어를 쪼개서, 그 중에 어떤 부분이 파일에 정의되어있는지 체크해서
그걸 argv로 만들어줘야한다. 
  nmt_parser = argparse.ArgumentParser()
  add_arguments(nmt_parser)
  FLAGS, unparsed = nmt_parser.parse_known_args()
  tf.app.run(main=main, argv=[sys.argv[0]] + unparsed)        #https://stackoverflow.com/questions/33703624/how-does-tf-app-run-work
보통 위처럼 세줄정도는 나올텐데, 그것보다는 한줄로 써서 위의 기능들을 취하는 것 아닐까?
그리고 main()으로 구분하면 가독성도 좋아지고, __main__이 아니어도 어떻게 접근이 가능할수도 있고..여러가지 가능성이 나오게되는 장점도 있다
"""
if __name__ == '__main__':
    # print(callable(main))
    # print(locals())
    tf.app.run()
