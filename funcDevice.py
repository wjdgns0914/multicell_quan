import tensorflow as tf
import numpy as np
import funcCustom
from openpyxl import load_workbook
FLAGS = tf.app.flags.FLAGS
# Load parameters setting
W_level=FLAGS.W_target_level
Wq_level=FLAGS.Wq_target_level
if W_level<2**10:
    choice=FLAGS.choice
    Inter=eval('['+FLAGS.Inter_variation_options+']')
    Intra=eval('['+FLAGS.Intra_variation_options+']')

    # Load device properties from the excel file
    # 1)Sigmoid functon
    mean_a = 3.156e+06 * 0.9996
    mean_b = 4.463 * 1.0647
    mean_c = 2.713 * 1.0004
    mean_d = -2.509e+04 * 1.0368
    wb = load_workbook(filename=FLAGS.level_info_name)
    sheetname = str(W_level) + 'Levels'
    if (Intra[0] or Inter[0]) and (str(wb.sheetnames).find(sheetname) > -1):
        try:
            ws = wb[sheetname]
            print("We will load writing information from the sheet %s" % sheetname)
        except:
            print("There is no sheet named as %s" % sheetname)
            exit()
        assert ws['A' + str(4 * choice + 1)] != None,"This sheet doesn't have voltage information"
        volt = np.array([cell.value for cell in tuple(ws.rows)[4 * choice]],dtype=np.float32)
        # r = np.array([cell.value for cell in tuple(ws.rows)[3 * choice+1]],dtype=np.float32)
        r_std = np.array([cell.value for cell in tuple(ws.rows)[4 * choice + 2]],dtype=np.float32)
        # r_ref = np.array([cell.value for cell in tuple(ws.rows)[3 * choice + 3][1:]],dtype=np.float32)
        r = mean_a / (1 + np.exp(-mean_b * (volt - mean_c))) - mean_d
        r_ref = (r[1:] + r[0:-1]) / 2
        # target resistance
        if Intra[0]:
            meanofstd = 0.6 * r_std
            stdofstd = 0.5 * meanofstd
            if Intra[1]:
                print("To be continued..Sorry!")
                exit(0)
            if Intra[3]:
                meanofstd = Intra[4] * meanofstd
                stdofstd = Intra[4] * stdofstd
        else:
            meanofstd = 0. * r_std
            stdofstd = 0. * meanofstd

        if Inter[0]:
            std_a = 3.156e+06 * 0.0698
            std_b = 4.463 * 0.1824
            std_c = 2.713 * 0.0354
            std_d = 2.509e+04 * 0.7708
            # stdofstd is free, 0일 수도 있고 아닐 수도 있다.
            if Inter[1]:
                print("To be continued..Sorry!")
                exit(0)
            if Inter[3]:
                std_a = Inter[4] * std_a
                std_b = Inter[4] * std_b
                std_c = Inter[4] * std_c
                std_d = Inter[4] * std_d
        else:
            std_a = 0.
            std_b = 0.
            std_c = 0.
            std_d = 0.
            stdofstd = 0. * meanofstd  # 무조건 0이어야한다.
    else:
        volt = np.linspace(1.5,4,W_level,dtype=np.float32)
        r = mean_a / (1 + np.exp(-mean_b * (volt - mean_c))) - mean_d
        r_std = np.zeros([W_level], dtype='float32')
        r_ref = (r[1:] + r[0:-1]) / 2
        std_a = 0.
        std_b = 0.
        std_c = 0.
        std_d = 0.
        meanofstd = 0. * r_std
        stdofstd = 0. * meanofstd
    write_info=[volt,r,r_ref]
    sigmoid_mean=[mean_a, mean_b, mean_c, mean_d]
    sigmoid_std=[std_a, std_b, std_c, std_d]
    file = open(FLAGS.checkpoint_dir + "/save_model.py", "a")
    funcCustom.magic_print("\n1.Hyper params setting")
    funcCustom.magic_print(
        "Dataset : ", FLAGS.dataset, "\nModel : ", FLAGS.model, "\nLR : ", FLAGS.learning_rate, "\nbatch_size : ", FLAGS.batch_size, file=file)
    funcCustom.magic_print(
        "W_level : ", FLAGS.W_target_level, "\nWq_level", FLAGS.Wq_target_level,"\nPatience : ", FLAGS.patience, file=file)
    funcCustom.magic_print("Inter=",Inter,file=file)
    funcCustom.magic_print("Intra=",Intra,file=file)
    funcCustom.magic_print(
        "Drift1 : ", FLAGS.Drift1, "\nDrift2 : ", FLAGS.Drift2, file=file)
    funcCustom.magic_print( '\n2.Write info',file=file)
    funcCustom.magic_print( ' volt=',volt,file=file)
    funcCustom.magic_print( '    r=',r,file=file)
    funcCustom.magic_print( 'r_std=',r_std,file=file)
    funcCustom.magic_print( 'r_ref=',r_ref,file=file)

    funcCustom.magic_print( '\n3.Cell info',file=file)
    funcCustom.magic_print( '[mean_a, mean_b, mean_c, mean_d]=',[mean_a,mean_b,mean_c,mean_d],file=file)
    funcCustom.magic_print( '[std_a , std_b , std_c , std_d ]=',[std_a,std_b,std_c,std_d],file=file)
    funcCustom.magic_print( '                       meanofstd=',meanofstd,file=file)
    funcCustom.magic_print( '                        stdofstd=\n',stdofstd,file=file)
    file.close()
else:
    write_info=[]
    sigmoid_mean=[0,0,0,0]
    sigmoid_std = [0, 0, 0, 0]

def generate_cell(array_shape,W_level=W_level):
    if W_level<2**10:
        a = tf.Variable(np.random.normal(loc=sigmoid_mean[0], scale=sigmoid_std[0], size=array_shape), trainable=False, dtype=tf.float32)
        b = tf.Variable(np.random.normal(loc=sigmoid_mean[1], scale=sigmoid_std[1], size=array_shape), trainable=False, dtype=tf.float32)
        c = tf.Variable(np.random.normal(loc=sigmoid_mean[2], scale=sigmoid_std[2], size=array_shape), trainable=False, dtype=tf.float32)
        d = tf.Variable(np.random.normal(loc=sigmoid_mean[3], scale=sigmoid_std[3], size=array_shape), trainable=False, dtype=tf.float32)
        std_val = tf.Variable(np.random.normal(loc=meanofstd, scale=stdofstd, size=array_shape + list(meanofstd.shape)),
                              trainable=False, dtype=tf.float32)
        cell_info=[a,b,c,d,std_val]
    else:
        cell_info=[]
    return write_info, cell_info